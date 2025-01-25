import requests
import openpyxl
from openpyxl.chart import LineChart, Reference
from openpyxl.chart.axis import DateAxis


#抓取台銀資料
def catch_data(url):
    datas = requests.get(url)   # 爬取網址內容
    datas.encoding = 'utf-8'    # 調整回應訊息編碼為 utf-8，避免編碼不同造成亂碼
    text = datas.text             # 以文字模式讀取內容
    rt = text.split('\n')[1:]   # 使用「換行」將內容拆分成串列 並跳過跳過第一行
    return rt
usd_rts = catch_data("https://rate.bot.com.tw/xrt/flcsv/0/L3M/USD")
jpy_rts = catch_data("https://rate.bot.com.tw/xrt/flcsv/0/L3M/JPY")


sale_rate = [['日期', '美金','日幣']]  # 建立串列，第一行為欄位名稱 
for i in range(len(usd_rts)):             
    try:                             # 使用 try 避開最後一行的空白行
        usd_daily = usd_rts[i].split(',')   
        jpy_daily = jpy_rts[i].split(',')


        year =usd_daily[0][0:4]
        month = usd_daily[0][4:6]
        day = usd_daily[0][6:8]
        date = year+"/"+month+"/"+day
        sale_rate.append([date ,float(usd_daily[13]),float(jpy_daily[13])])

    except:
      break

#匯入excel檔案
wb = openpyxl.Workbook()
ws = wb.active

for row in sale_rate:
       ws.append(row)

#繪製圖表

chart = LineChart()                
# chart.style = 15   
chart.title = '美金匯率'       
chart.y_axis.title = '匯率'    
chart.x_axis.title = '日期'        

usd_data = Reference(ws, min_col=2 , min_row=1,
                 max_row=ws.max_row)  

chart.add_data(usd_data, titles_from_data=True)  # 建立圖表
xtitle = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
chart.set_categories(xtitle)

chart.width = 35
chart.height = 15
ws.add_chart(chart, 'D1')      

chart2 = LineChart()                 
chart2.title = '日幣匯率'
chart2.y_axis.title = '匯率'
chart2.x_axis.title = '日期'

jpy_data = Reference(ws, min_col=3 , min_row=1, max_row=ws.max_row)
chart2.add_data(jpy_data, titles_from_data=True)
chart2.set_categories(xtitle)

chart2.width = 35
chart2.height = 15
ws.add_chart(chart2, 'D35')

wb.save("匯率及時更新.xlsx") 

